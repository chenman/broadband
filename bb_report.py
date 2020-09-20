# !usr/bin/env python
# -*- coding:utf-8 -*-
"""
@author:Chen Man
@file:  bb_report
@time:  2020/9/20
"""
import oss_login as oss

import datetime
import pandas as pd
import openpyxl


def copy_file(to_filename, from_filename):
    to_file = openpyxl.load_workbook(to_filename)
    to_sheet = to_file.worksheets[0]

    from_file = openpyxl.load_workbook(from_filename)
    from_sheet = from_file.worksheets[0]
    nrow = to_sheet.max_row
    ncol = to_sheet.max_column

    nrow2 = from_sheet.max_row

    for i in range(2, nrow2 + 1):
        for j in range(1, ncol + 1):
            to_sheet.cell(nrow+i-1, j).value = from_sheet.cell(i, j).value

    to_file.save(to_filename)


def export_accept_list(begin_time, end_time, cookies, filename):
    accept_request_xml = '<?xml version="1.0" encoding="gb2312"?><requestdata><parameter ' \
                         'name="reportCode">orderNotRealTime</parameter><parameter ' \
                         'name="ttOrgId">10000000</parameter><parameter name="priv">infoHide</parameter><parameter ' \
                         'name="areaIds">4,102,41,42,43,44,45</parameter><parameter ' \
                         'name="serviceIds">220323,220372</parameter><parameter ' \
                         'name="searchType">OrdMoni</parameter><parameter ' \
                         'name="startDate">%s</parameter><parameter name="endDate">%s</parameter><parameter ' \
                         'name="DateType">1</parameter><parameter ' \
                         'name="searchFlag">true</parameter><parameter name="staffId">223214</parameter><parameter ' \
                         'name="pageIndex">1</parameter><parameter name="pageSize">10000</parameter></requestdata> ' \
                         % (begin_time.strftime('%Y-%m-%d %H:%M:%S'), end_time.strftime('%Y-%m-%d %H:%M:%S'))
    oss.export_data(accept_request_xml, cookies, filename=filename)


def export_finish_list(begin_time, end_time, cookies, filename):
    finish_request_xml = '<?xml version="1.0" encoding="gb2312"?><requestdata><parameter ' \
                         'name="reportCode">orderNotRealTime</parameter><parameter ' \
                         'name="ttOrgId">10000000</parameter><parameter name="priv">infoHide</parameter><parameter ' \
                         'name="areaIds">4,102,41,42,43,44,45</parameter><parameter ' \
                         'name="serviceIds">220323</parameter><parameter name="omStates">10F</parameter><parameter ' \
                         'name="searchType">OrdMoni</parameter><parameter name="startDate">%s</parameter><parameter ' \
                         'name="endDate">%s</parameter><parameter ' \
                         'name="DateType">2</parameter><parameter name="searchFlag">true</parameter><parameter ' \
                         'name="staffId">223214</parameter><parameter name="pageIndex">1</parameter><parameter ' \
                         'name="pageSize">30000</parameter></requestdata>' \
                         % (begin_time.strftime('%Y-%m-%d %H:%M:%S'), end_time.strftime('%Y-%m-%d %H:%M:%S'))
    oss.export_data(finish_request_xml, cookies, filename=filename)


def export_accept_finish_list(now, cookies):
    current_month_first_day = datetime.date(year=now.year, month=now.month, day=1)
    month_10_day = datetime.datetime(year=now.year, month=now.month, day=10)
    month_20_day = datetime.datetime(year=now.year, month=now.month, day=20)
    accept_list_file = now.strftime("%Y%m%d%H%M%S-accept") + '.xlsx'
    finish_list_file = now.strftime("%Y%m%d%H%M%S-finish") + '.xlsx'
    if month_10_day > now:
        export_accept_list(current_month_first_day, now, cookies, accept_list_file)
        export_finish_list(current_month_first_day, now, cookies, finish_list_file)
    elif month_10_day <= now < month_20_day:
        export_accept_list(current_month_first_day, month_10_day - datetime.timedelta(seconds=1), cookies, accept_list_file)
        export_finish_list(current_month_first_day, month_10_day - datetime.timedelta(seconds=1), cookies, finish_list_file)
        accept_tmp1_file = now.strftime("%Y%m%d%H%M%S-accept_1") + '.xlsx'
        finish_tmp1_file = now.strftime("%Y%m%d%H%M%S-finish_1") + '.xlsx'
        export_accept_list(month_10_day, now, cookies, accept_tmp1_file)
        export_finish_list(month_10_day, now, cookies, finish_tmp1_file)
        copy_file(accept_list_file, accept_tmp1_file)
        copy_file(finish_list_file, finish_tmp1_file)
    else:
        export_accept_list(current_month_first_day, month_10_day - datetime.timedelta(seconds=1), cookies, accept_list_file)
        export_finish_list(current_month_first_day, month_10_day - datetime.timedelta(seconds=1), cookies, finish_list_file)
        accept_tmp1_file = now.strftime("%Y%m%d%H%M%S-accept_1") + '.xlsx'
        finish_tmp1_file = now.strftime("%Y%m%d%H%M%S-finish_1") + '.xlsx'
        export_accept_list(month_10_day, month_20_day - datetime.timedelta(seconds=1), cookies, accept_tmp1_file)
        export_finish_list(month_10_day, month_20_day - datetime.timedelta(seconds=1), cookies, finish_tmp1_file)
        copy_file(accept_list_file, accept_tmp1_file)
        copy_file(finish_list_file, finish_tmp1_file)
        accept_tmp2_file = now.strftime("%Y%m%d%H%M%S-accept_2") + '.xlsx'
        finish_tmp2_file = now.strftime("%Y%m%d%H%M%S-finish_2") + '.xlsx'
        export_accept_list(month_20_day, now, cookies, accept_tmp2_file)
        export_finish_list(month_20_day, now, cookies, finish_tmp2_file)
        copy_file(accept_list_file, accept_tmp2_file)
        copy_file(finish_list_file, finish_tmp2_file)


if __name__ == '__main__':
    # print(int(time.time() * 1000))
    cookies = oss.login('chenman', '!EIje*62')
    now = datetime.datetime.now()
    export_accept_finish_list(now, cookies)
    init_date = datetime.date.today()
    pre_date = init_date - datetime.timedelta(days=1)
    current_month_first_day = datetime.date(year=now.year, month=now.month, day=1)

    accept_list_file = now.strftime("%Y%m%d%H%M%S-accept") + '.xlsx'
    finish_list_file = now.strftime("%Y%m%d%H%M%S-finish") + '.xlsx'

    cfg_org = pd.read_excel("config.xlsx", sheet_name="sheet1")
    cfg_grd = pd.read_excel("config.xlsx", sheet_name="sheet2")
    cfg_mgr = pd.read_excel("config.xlsx", sheet_name="sheet3")

    # print(cfg_org)
    # print(cfg_grd)
    # print(cfg_mgr)

    accept_list = pd.read_excel(accept_list_file, '全量工单信息')

    accept_list_2d = accept_list[accept_list['受理时间'] >= pre_date.strftime("%Y-%m-%d %H:%M:%S")]

    al = pd.merge(accept_list_2d, cfg_org, how="left", on="受理部门ID")  # 比对网格长、网格经理（受理清单）

    # 计算今日受理量
    day_accept = al[al['受理时间'] >= init_date.strftime("%Y-%m-%d %H:%M:%S")].groupby(al['网格长']).agg(
        {'工单编号': 'count'}).rename(
        columns={'工单编号': '网点当日受理量'})  # 计算网格长当日受理量（受理清单）
    day_accept = pd.DataFrame(day_accept)
    day_accept.reset_index(inplace=True)
    day_accept = pd.merge(cfg_mgr, day_accept, on="网格长", how="outer").fillna(0)
    day_accept['网点当日受理量'] = day_accept['网点当日受理量'].apply(int)  # 今日受理量
    # print(day_accept)

    # 计算昨日受理量
    pre_accept = al[(al['受理时间'] >= pre_date.strftime("%Y-%m-%d %H:%M:%S")) & (
            al['受理时间'] < init_date.strftime("%Y-%m-%d %H:%M:%S"))].groupby(al['网格长']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '网点昨日受理量'})  # 计算网格长当日受理量（受理清单）
    pre_accept = pd.DataFrame(pre_accept)
    pre_accept.reset_index(inplace=True)
    pre_accept = pd.merge(cfg_mgr, pre_accept, on="网格长", how="outer").fillna(0)
    pre_accept['网点昨日受理量'] = pre_accept['网点昨日受理量'].apply(int)  # 昨日受理量
    # print(pre_accept)

    # 竣工信息
    finish_list = pd.read_excel(finish_list_file, '全量工单信息')

    # 网点竣工信息
    ofl = pd.merge(finish_list, cfg_org, how="left", on="受理部门ID")  # 比对网格长、网格经理（受理清单）

    # 网点当月竣工量
    o_month_finish = ofl.groupby(ofl['网格长']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '网点当月竣工量'})
    o_month_finish = pd.DataFrame(o_month_finish)
    o_month_finish.reset_index(inplace=True)
    o_month_finish = pd.merge(cfg_mgr, o_month_finish, on="网格长", how="outer").fillna(0)
    o_month_finish['网点当月竣工量'] = o_month_finish['网点当月竣工量'].apply(int)  # 昨日受理量
    # print(o_month_finish)

    # 网格当月竣工量
    gfl = pd.merge(finish_list, cfg_grd, how="left", on="所属网格")

    # 网格当月竣工量
    g_month_finish = gfl.groupby(gfl['网格长']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '网格当月竣工量'})
    g_month_finish = pd.DataFrame(g_month_finish)
    g_month_finish.reset_index(inplace=True)
    g_month_finish = pd.merge(cfg_mgr, g_month_finish, on="网格长", how="outer").fillna(0)
    g_month_finish['网格当月竣工量'] = g_month_finish['网格当月竣工量'].apply(int)  # 昨日受理量
    # print(g_month_finish)

    # 网格受理信息
    gal = pd.merge(accept_list_2d, cfg_grd, how="left", on="所属网格")  # 比对网格长、网格经理（受理清单）

    # 计算网格今日受理量
    g_day_accept = gal[gal['受理时间'] >= init_date.strftime("%Y-%m-%d %H:%M:%S")].groupby(gal['网格长']).agg(
        {'工单编号': 'count'}).rename(
        columns={'工单编号': '网格当日受理量'})  # 计算网格长当日受理量（受理清单）
    g_day_accept = pd.DataFrame(g_day_accept)
    g_day_accept.reset_index(inplace=True)
    g_day_accept = pd.merge(cfg_mgr, g_day_accept, on="网格长", how="outer").fillna(0)
    g_day_accept['网格当日受理量'] = g_day_accept['网格当日受理量'].apply(int)  # 今日受理量
    # print(g_day_accept)

    # 计算网格昨日受理量
    g_pre_accept = gal[(gal['受理时间'] >= pre_date.strftime("%Y-%m-%d %H:%M:%S")) & (
            gal['受理时间'] < init_date.strftime("%Y-%m-%d %H:%M:%S"))].groupby(gal['网格长']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '网格昨日受理量'})
    g_pre_accept = pd.DataFrame(g_pre_accept)
    g_pre_accept.reset_index(inplace=True)
    g_pre_accept = pd.merge(cfg_mgr, g_pre_accept, on="网格长", how="outer").fillna(0)
    g_pre_accept['网格昨日受理量'] = g_pre_accept['网格昨日受理量'].apply(int)  # 昨日受理量
    # print(g_pre_accept)

    result = pd.merge(o_month_finish, day_accept, on='网格长', how='left')
    result = pd.merge(result, pre_accept, on='网格长', how='left')
    result = pd.merge(result, g_month_finish, on='网格长', how='left')
    result = pd.merge(result, g_day_accept, on='网格长', how='left')
    result = pd.merge(result, g_pre_accept, on='网格长', how='left')
    # 保存网格报表
    result.to_excel("grid_static.xlsx", index=True, sheet_name='网格')

    """
    区县受理及竣工统计报表
    """
    month_accept = accept_list.groupby(accept_list['区县']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '区县当月受理量'})  # 区县月受理量
    month_accept = pd.DataFrame(month_accept)
    month_accept.reset_index(inplace=True)

    today_accept = accept_list[accept_list['受理时间'] >= init_date.strftime("%Y-%m-%d %H:%M:%S")] \
        .groupby(accept_list['区县']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '区县当日受理'})
    today_accept = pd.DataFrame(today_accept)
    today_accept.reset_index(inplace=True)

    month_finish = finish_list.groupby(finish_list['区县']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '区县当月竣工量'})  # 区县月竣工量
    month_finish = pd.DataFrame(month_finish)
    month_finish.reset_index(inplace=True)

    today_finish = finish_list[finish_list['竣工时间'] >= init_date.strftime("%Y-%m-%d %H:%M:%S")] \
        .groupby(finish_list['区县']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '区县当日竣工'})
    today_finish = pd.DataFrame(today_finish)
    today_finish.reset_index(inplace=True)

    result = pd.merge(month_accept, today_accept, on='区县', how='left')
    result = pd.merge(result, month_finish, on='区县', how='left')
    result = pd.merge(result, today_finish, on='区县', how='left')
    result.to_excel("county_static.xlsx", index=True, sheet_name='区县')

    """
    渠道网点受理及竣工统计报表
    受理部门 截止昨日月竣工	今日受理	昨日受理
    """
    org_day_accept = accept_list_2d[(accept_list_2d['受理时间'] >= init_date.strftime("%Y-%m-%d %H:%M:%S"))] \
        .groupby(accept_list_2d['受理部门']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '网点今日受理'})
    org_day_accept = pd.DataFrame(org_day_accept)
    org_day_accept.reset_index(inplace=True)

    org_pre_accept = accept_list_2d[(accept_list_2d['受理时间'] >= pre_date.strftime("%Y-%m-%d %H:%M:%S"))
                                    & (accept_list_2d['受理时间'] < init_date.strftime("%Y-%m-%d %H:%M:%S"))] \
        .groupby(accept_list_2d['受理部门']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '网点昨日受理'})
    org_pre_accept = pd.DataFrame(org_pre_accept)
    org_pre_accept.reset_index(inplace=True)

    org_total_finish = finish_list[(finish_list['竣工时间'] < init_date.strftime("%Y-%m-%d %H:%M:%S"))] \
        .groupby(finish_list['受理部门']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '网点截止昨日竣工量'})
    org_total_finish = pd.DataFrame(org_total_finish)
    org_total_finish.reset_index(inplace=True)
    result = pd.merge(org_total_finish, org_day_accept, on='受理部门', how='outer').fillna(0)
    result = pd.merge(result, org_pre_accept, on='受理部门', how='outer').fillna(0)
    result.to_excel('org_static.xlsx', index=True, sheet_name='网点')

    """
    铁通网点受理及竣工统计报表
    """
    tt_today_accept = accept_list_2d[(accept_list_2d['受理部门'].str.contains('铁通')) &
                                     (accept_list_2d['受理时间'] >= init_date.strftime("%Y-%m-%d %H:%M:%S"))]\
        .groupby(accept_list_2d['受理部门']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '铁通当日受理量'})
    tt_today_accept = pd.DataFrame(tt_today_accept)
    tt_today_accept.reset_index(inplace=True)
    
    tt_total_finish = finish_list[(finish_list['受理部门'].str.contains('铁通')) &
                                  (finish_list['竣工时间'] < init_date.strftime("%Y-%m-%d %H:%M:%S"))]\
        .groupby(finish_list['受理部门']).agg({'工单编号': 'count'}).rename(
        columns={'工单编号': '铁通截止昨日竣工量'})
    tt_total_finish = pd.DataFrame(tt_total_finish)
    tt_total_finish.reset_index(inplace=True)
    
    result = pd.merge(tt_today_accept, tt_total_finish, on='受理部门', how='outer').fillna(0)
    result.to_excel('tietong_static.xlsx', index=True, sheet_name='铁通')
