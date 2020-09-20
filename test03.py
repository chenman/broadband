# !usr/bin/env python
# -*- coding:utf-8 -*-
"""
@author:Chen Man
@file:  test03
@time:  2020/9/20
"""

import datetime
import oss_login as oss
import openpyxl


def copy_file(to_filename, from_filename):
    to_file = openpyxl.load_workbook(to_filename)
    to_sheet = to_file.worksheets[0]

    from_file = openpyxl.load_workbook(from_filename)
    from_sheet = from_file.worksheets[0]
    nrow = to_sheet.max_row
    ncol = to_sheet.max_column

    nrow2 = from_sheet.max_row

    for i in range(2, nrow2 + 1):
        for j in range(1, ncol + 1):
            to_sheet.cell(nrow+i-1, j).value = from_sheet.cell(i, j).value

    to_file.save(to_filename)


def export_accept_list(begin_time, end_time, cookies, filename):
    accept_request_xml = '<?xml version="1.0" encoding="gb2312"?><requestdata><parameter ' \
                         'name="reportCode">orderNotRealTime</parameter><parameter ' \
                         'name="ttOrgId">10000000</parameter><parameter name="priv">infoHide</parameter><parameter ' \
                         'name="areaIds">4,102,41,42,43,44,45</parameter><parameter ' \
                         'name="serviceIds">220323,220372</parameter><parameter ' \
                         'name="searchType">OrdMoni</parameter><parameter ' \
                         'name="startDate">%s</parameter><parameter name="endDate">%s</parameter><parameter ' \
                         'name="DateType">1</parameter><parameter ' \
                         'name="searchFlag">true</parameter><parameter name="staffId">223214</parameter><parameter ' \
                         'name="pageIndex">1</parameter><parameter name="pageSize">10000</parameter></requestdata> ' \
                         % (begin_time.strftime('%Y-%m-%d %H:%M:%S'), end_time.strftime('%Y-%m-%d %H:%M:%S'))
    oss.export_data(accept_request_xml, cookies, filename=filename)


def export_finish_list(begin_time, end_time, cookies, filename):
    finish_request_xml = '<?xml version="1.0" encoding="gb2312"?><requestdata><parameter ' \
                         'name="reportCode">orderNotRealTime</parameter><parameter ' \
                         'name="ttOrgId">10000000</parameter><parameter name="priv">infoHide</parameter><parameter ' \
                         'name="areaIds">4,102,41,42,43,44,45</parameter><parameter ' \
                         'name="serviceIds">220323</parameter><parameter name="omStates">10F</parameter><parameter ' \
                         'name="searchType">OrdMoni</parameter><parameter name="startDate">%s</parameter><parameter ' \
                         'name="endDate">%s</parameter><parameter ' \
                         'name="DateType">2</parameter><parameter name="searchFlag">true</parameter><parameter ' \
                         'name="staffId">223214</parameter><parameter name="pageIndex">1</parameter><parameter ' \
                         'name="pageSize">30000</parameter></requestdata>' \
                         % (begin_time.strftime('%Y-%m-%d %H:%M:%S'), end_time.strftime('%Y-%m-%d %H:%M:%S'))
    oss.export_data(finish_request_xml, cookies, filename=filename)


def export_accept_finish_list(now, cookies):
    current_month_first_day = datetime.date(year=now.year, month=now.month, day=1)
    month_10_day = datetime.datetime(year=now.year, month=now.month, day=10)
    month_20_day = datetime.datetime(year=now.year, month=now.month, day=20)
    accept_list_file = now.strftime("%Y%m%d%H%M%S-accept") + '.xlsx'
    finish_list_file = now.strftime("%Y%m%d%H%M%S-finish") + '.xlsx'
    if month_10_day > now:
        export_accept_list(current_month_first_day, now, accept_list_file)
        export_finish_list(current_month_first_day, now, finish_list_file)
    elif month_10_day <= now < month_20_day:
        export_accept_list(current_month_first_day, month_10_day - datetime.timedelta(seconds=1), accept_list_file)
        export_finish_list(current_month_first_day, month_10_day - datetime.timedelta(seconds=1), finish_list_file)
        accept_tmp1_file = now.strftime("%Y%m%d%H%M%S-accept_1") + '.xlsx'
        finish_tmp1_file = now.strftime("%Y%m%d%H%M%S-finish_1") + '.xlsx'
        export_accept_list(month_10_day, now, accept_tmp1_file)
        export_finish_list(month_10_day, now, finish_tmp1_file)
        copy_file(accept_list_file, accept_tmp1_file)
        copy_file(finish_list_file, finish_tmp1_file)
    else:
        export_accept_list(current_month_first_day, month_10_day - datetime.timedelta(seconds=1), accept_list_file)
        export_finish_list(current_month_first_day, month_10_day - datetime.timedelta(seconds=1), finish_list_file)
        accept_tmp1_file = now.strftime("%Y%m%d%H%M%S-accept_1") + '.xlsx'
        finish_tmp1_file = now.strftime("%Y%m%d%H%M%S-finish_1") + '.xlsx'
        export_accept_list(month_10_day, month_20_day - datetime.timedelta(seconds=1), accept_tmp1_file)
        export_finish_list(month_10_day, month_20_day - datetime.timedelta(seconds=1), finish_tmp1_file)
        copy_file(accept_list_file, accept_tmp1_file)
        copy_file(finish_list_file, finish_tmp1_file)
        accept_tmp2_file = now.strftime("%Y%m%d%H%M%S-accept_2") + '.xlsx'
        finish_tmp2_file = now.strftime("%Y%m%d%H%M%S-finish_2") + '.xlsx'
        export_accept_list(month_20_day, now, accept_tmp2_file)
        export_finish_list(month_20_day, now, finish_tmp2_file)
        copy_file(accept_list_file, accept_tmp2_file)
        copy_file(finish_list_file, finish_tmp2_file)



if __name__ == '__main__':
    copy_file("tmp01.xlsx", "tmp02.xlsx")
    # print(int(time.time() * 1000))
    # now = datetime.datetime.now()
    # init_date = datetime.date.today()
    # pre_date = init_date - datetime.timedelta(days=1)
    # current_month_first_day = datetime.date(year=now.year, month=now.month, day=1)
    #
    # cookies = oss.login('chenman', '!EIje*62')
    # accept_request_xml = '<?xml version="1.0" encoding="gb2312"?><requestdata><parameter ' \
    #                      'name="reportCode">orderNotRealTime</parameter><parameter ' \
    #                      'name="ttOrgId">10000000</parameter><parameter name="priv">infoHide</parameter><parameter ' \
    #                      'name="areaIds">4,102,41,42,43,44,45</parameter><parameter ' \
    #                      'name="serviceIds">220323,220372</parameter><parameter ' \
    #                      'name="searchType">OrdMoni</parameter><parameter ' \
    #                      'name="startDate">%s</parameter><parameter name="endDate">%s</parameter><parameter ' \
    #                      'name="DateType">1</parameter><parameter ' \
    #                      'name="searchFlag">true</parameter><parameter name="staffId">223214</parameter><parameter ' \
    #                      'name="pageIndex">1</parameter><parameter name="pageSize">20000</parameter></requestdata> ' \
    #                      % (current_month_first_day.strftime('%Y-%m-%d %H:%M:%S'), now.strftime('%Y-%m-%d %H:%M:%S'))
    #
    # accept_list_file = now.strftime("%Y%m%d%H%M%S-accept") + '.xlsx'
    # oss.export_data(accept_request_xml, cookies, filename=accept_list_file)